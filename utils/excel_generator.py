"""
Excel generation using openpyxl — produces the instrumentation .xlsx file
matching the company's exact format.
"""
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# -- Style constants --
FONT_DEFAULT = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_HYPERLINK = Font(name="Arial", size=10, color="0563C1", underline="single")
FONT_NEW = Font(name="Arial", size=10, color="0070C0")
FONT_UPDATE = Font(name="Arial", size=10, color="FF6600")
FONT_EXISTS = Font(name="Arial", size=10)

FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FILL_STORY = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALIGNMENT = Alignment(wrap_text=True, vertical="top")

COLUMN_WIDTHS = {
    "A": 22,
    "B": 33,
    "C": 30,
    "D": 90,
    "E": 15,
    "F": 18,
    "G": 14,
    "H": 30,
    "I": 30,
    "J": 15,
}

HEADERS = [
    "Story",
    "Name",
    "Trigger",
    "Event Specific Payload",
    "Common Payload",
    "Event Status",
    "AAT + Priority",
    "Notes",
    "Metrics",
    "Can Be Tracked",
]

# Windows/macOS reserved characters; also strip Excel-invalid sheet name chars like [ ].
_INVALID_BASENAME_CHARS = re.compile(r'[<>:"/\\|?*\[\]\x00-\x1f]')


def sanitize_page_basename(raw: str | None) -> str:
    """
    Safe base name for download filenames and workbook metadata.
    If empty after trim, use a timestamped fallback so back-to-back exports do not collide.
    """
    if raw is None:
        raw = ""
    s = _INVALID_BASENAME_CHARS.sub("_", str(raw).strip())
    s = s.strip(" .")
    if not s:
        return f"instrumentation_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    return s[:200]


def generate_excel(
    rows: list[dict],
    page_name: str,
    figma_url: str = "",
) -> io.BytesIO:
    """
    Generate an .xlsx BytesIO buffer from instrumentation rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = page_name[:31] if page_name else "Instrumentation"

    # -- Set column widths --
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # -- Row 1: PRD --
    ws["A1"] = "PRD"
    ws["A1"].font = FONT_BOLD
    ws["B1"] = f"{page_name} PRD"
    ws["B1"].font = FONT_DEFAULT

    # -- Row 2: Figma --
    ws["A2"] = "Figma"
    ws["A2"].font = FONT_BOLD
    if figma_url:
        ws["B2"] = figma_url
        ws["B2"].font = FONT_HYPERLINK
        ws["B2"].hyperlink = figma_url
    else:
        ws["B2"] = ""

    # -- Row 3: Other Docs --
    ws["A3"] = "Other Docs"
    ws["A3"].font = FONT_BOLD

    # -- Row 4: blank --

    # -- Row 5: Headers --
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGNMENT
        cell.border = THIN_BORDER

    # -- Data rows (row 6+) --
    prev_story = None
    for row_idx, row_data in enumerate(rows, start=6):
        story = row_data.get("story", "")
        name = row_data.get("name", "")
        trigger = row_data.get("trigger", "")
        payload = row_data.get("event_specific_payload", "")
        common = row_data.get("common_payload", "No Change")
        status = row_data.get("event_status", "New")
        priority = row_data.get("aat_priority", "P2")
        notes = row_data.get("notes", "")
        metrics = row_data.get("metrics", "")

        # Story continuation logic
        show_story = story if story != prev_story else ""
        prev_story = story

        values = [show_story, name, trigger, payload, common, status, priority, notes, metrics, "Yes"]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_DEFAULT
            cell.alignment = ALIGNMENT
            cell.border = THIN_BORDER

            # Story fill
            if col_idx == 1 and show_story:
                cell.fill = FILL_STORY

            # Event Status color
            if col_idx == 6:
                if status == "New":
                    cell.font = FONT_NEW
                elif status == "Exists - Update":
                    cell.font = FONT_UPDATE
                else:
                    cell.font = FONT_EXISTS

    # -- Freeze pane below headers --
    ws.freeze_panes = "A6"

    # -- Apply default font to metadata rows --
    for r in range(1, 5):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if not cell.font or cell.font == Font():
                cell.font = FONT_DEFAULT

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
